"""
LoL Stats - Excel Importer
============================
Parses the Coach K "Support champ guide" spreadsheet into the
3-context strategy data model:

    vs_support   - how to play VS an enemy support   (sheets: "Support Counters", "sc bu")
    with_jungler - synergy WITH your allied jungler   (sheet: "Jungle Synergy")
    with_adc     - laning WITH your allied ADC        (sheet: "Laning Guide")

Each sheet uses merged cells + multi-row champion blocks, so we parse
them with explicit column maps rather than generic header detection.
"""

import logging
import re
from pathlib import Path
from typing import Optional

import openpyxl

logger = logging.getLogger(__name__)

DEFAULT_EXCEL_PATH = Path(__file__).parent.parent.parent / "Support champ guide.xlsx"

# Title row is row 1, header row is row 2 → data starts at row 3 (1-indexed).
DATA_START_ROW = 3

# ── Per-sheet column maps (0-indexed, validated by inspection) ──
SHEET_CONFIG = {
    # "Support Counters" + "sc bu" share the same layout
    "support_counters": {
        "sheets": ["Support Counters", "sc bu"],
        "context": "vs_support",
        "name_col": 1,
        "scalar_cols": {
            "early_strength": 2,
            "more_info": 13,
            "tier": 14,
            "adc_synergy": 15,
        },
        "multiline_cols": {  # cell text → list of bullets
            "strengths": 3,
            "weaknesses": 4,
            "how_to_play": 11,
        },
        "list_cols": [6, 8, 10],  # champion-name list, gathered across the block rows
    },
    "jungle_synergy": {
        "sheets": ["Jungle Synergy"],
        "context": "with_jungler",
        "name_col": 1,
        "scalar_cols": {
            "early": 2,
            "pathing": 3,
            "synergy": 4,
            "vision_level1": 15,
            "gameplan": 16,
            "important_info": 17,
        },
        "multiline_cols": {},
        "list_cols": [6, 8, 10, 12, 14],
    },
    "laning_guide": {
        "sheets": ["Laning Guide"],
        "context": "with_adc",
        "name_col": 1,
        "scalar_cols": {
            "strength": 2,
            "synergy": 3,
            "gameplan": 14,
            "how_to_trade": 15,
            "when_to_roam": 16,
        },
        "multiline_cols": {},
        "list_cols": [5, 7, 9, 11, 13],
    },
}


def import_from_excel(excel_path: Optional[str] = None) -> dict:
    """
    Parse the Coach K spreadsheet and return champion entries in the
    3-context strategy format.

    Returns:
        {"entries": {champ: {context: {...}}}, "errors": [...]}
    """
    path = Path(excel_path) if excel_path else DEFAULT_EXCEL_PATH

    if not path.exists():
        raise FileNotFoundError(f"Excel file not found: {path}")

    logger.info("Parsing Excel file: %s", path)

    try:
        wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    except Exception as e:
        raise ValueError(f"Failed to open Excel file: {e}")

    entries: dict[str, dict] = {}
    errors: list[str] = []
    available = set(wb.sheetnames)

    for cfg in SHEET_CONFIG.values():
        for sheet_name in cfg["sheets"]:
            if sheet_name not in available:
                errors.append(f"Sheet '{sheet_name}' not found in workbook.")
                continue
            try:
                _parse_structured_sheet(wb[sheet_name], cfg, entries)
            except Exception as e:
                logger.warning("Failed to parse sheet '%s': %s", sheet_name, e)
                errors.append(f"Sheet '{sheet_name}': {e}")

    wb.close()

    if not entries:
        errors.append("No champion entries could be extracted. Check file format.")
        logger.warning("No champion entries extracted from Excel file.")

    return {"entries": entries, "errors": errors}


def _parse_structured_sheet(sheet, cfg: dict, entries: dict) -> None:
    """Parse one sheet using its column config, merging into `entries`."""
    name_col = cfg["name_col"]
    context = cfg["context"]

    # Read all non-empty rows up front (read_only mode iterates once).
    rows = []
    for row in sheet.iter_rows(min_row=DATA_START_ROW, values_only=True):
        if any(c is not None and str(c).strip() for c in row):
            rows.append(list(row))

    # Group rows into champion blocks. A new block starts whenever the
    # name column is non-empty; subsequent rows (blank name) belong to it.
    blocks: list[tuple[str, list[list]]] = []
    current_name = None
    current_rows: list[list] = []

    for row in rows:
        name = _cell(row, name_col)
        if name:
            if current_name:
                blocks.append((current_name, current_rows))
            current_name = name
            current_rows = [row]
        elif current_name:
            current_rows.append(row)

    if current_name:
        blocks.append((current_name, current_rows))

    for raw_name, block_rows in blocks:
        clean = _clean_champion_name(raw_name)
        if not clean:
            continue

        ctx_data = _extract_context(block_rows, cfg)
        if not ctx_data:
            continue

        champ = entries.setdefault(clean, {})
        # Merge: if the same champ already has this context (e.g. from sc bu +
        # Support Counters), combine list fields without duplicates.
        if context in champ:
            _merge_context(champ[context], ctx_data)
        else:
            champ[context] = ctx_data


def _extract_context(block_rows: list[list], cfg: dict) -> dict:
    """Extract one context dict from a champion's row block."""
    first = block_rows[0] if block_rows else []
    out: dict = {}

    # Scalar text fields (from the first row only).
    for key, col in cfg["scalar_cols"].items():
        val = _cell(first, col)
        if val:
            out[key] = val

    # Multiline → bullet list (from the first row).
    for key, col in cfg["multiline_cols"].items():
        val = _cell(first, col)
        items = _split_bullets(val)
        if items:
            out[key] = items

    # Champion-name lists gathered across ALL rows in the block.
    if cfg["list_cols"]:
        names: list[str] = []
        for row in block_rows:
            for col in cfg["list_cols"]:
                val = _cell(row, col)
                if val:
                    cleaned = _clean_champion_name(val)
                    if cleaned and cleaned not in names:
                        names.append(cleaned)
        if names:
            # "counters" for vs_support, "best_supports" for the synergy sheets
            list_key = "counters" if cfg["context"] == "vs_support" else "best_supports"
            out[list_key] = names

    return out


# ── Cell helpers ──────────────────────────────────────────────

def _cell(row: list, col: int) -> str:
    """Return stripped string value of row[col], or '' if missing."""
    if col >= len(row):
        return ""
    v = row[col]
    if v is None:
        return ""
    s = str(v).strip()
    # Skip Excel error sentinels.
    if s.startswith("#") and s.endswith("?"):
        return ""
    return s


def _split_bullets(text: str) -> list[str]:
    """Split a multiline cell into cleaned bullet items."""
    if not text:
        return []
    items = []
    for raw in re.split(r"[\n\r]+", text):
        line = raw.strip()
        if not line:
            continue
        # Strip a leading bullet marker (-, •, *, –, ·)
        line = re.sub(r"^[-•*\–·]+\s*", "", line).strip()
        if line:
            items.append(line)
    return items


def _clean_champion_name(raw: str) -> str:
    """Normalize a champion name from an Excel cell."""
    raw = (raw or "").strip()
    if not raw:
        return ""
    # Drop emoji/symbols but keep letters, digits, apostrophes, hyphens, dots, spaces.
    raw = re.sub(r"[^\w\s'\-\.]", "", raw)
    raw = re.sub(r"\s+", " ", raw).strip()
    if len(raw) < 2:
        return ""
    # Capitalize the first letter only (preserves "Bel'Veth", "Dr. Mundo").
    return raw[0].upper() + raw[1:]


def _merge_context(target: dict, source: dict) -> None:
    """Merge two context blocks: combine lists, fill missing scalars."""
    for key, val in source.items():
        if isinstance(val, list):
            existing = target.setdefault(key, [])
            for item in val:
                if item not in existing:
                    existing.append(item)
        elif val and not target.get(key):
            target[key] = val
